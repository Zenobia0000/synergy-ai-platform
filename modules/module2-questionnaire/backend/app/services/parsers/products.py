"""
產品目錄與圖片連結解析器。

主要函式：parse_products
"""

from __future__ import annotations

import difflib
import hashlib
import logging
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
from docx import Document

from .types import Meta, Product, ProductsSchema, UnmatchedImage

logger = logging.getLogger(__name__)

PARSER_VERSION = "0.1.0"

# 價格抽取正則：抓第一組數字（含逗號）
_PRICE_RE = re.compile(r"[\d,]+")

# 適用情境分隔符
_SCENARIO_SPLIT_RE = re.compile(r"[、,；;\n]")

# 圖片 docx 解析正則：支援中英文冒號、可選 [ 前綴
_IMAGE_RE = re.compile(
    r"^(?P<name>.+?)[:：]\s*\[?(?P<url>https?://\S+?)[\]>]?\s*$"
)

# 清理 docx 名稱：去括號內容、去尾端符號
_BRACKET_RE = re.compile(r"[（(（][^)）)]*[)）)]")
_TRAILING_RE = re.compile(r"[-\s]+$")

FUZZY_THRESHOLD = 0.85


def _sha256_file(path: Path) -> str:
    """計算檔案 SHA256 hex digest。"""
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _normalize_name(name: str) -> str:
    """正規化產品名稱：NFKC + 去括號 + 去尾端符號。"""
    name = unicodedata.normalize("NFKC", name)
    name = _BRACKET_RE.sub("", name)
    name = _TRAILING_RE.sub("", name.strip())
    return name.strip()


# 去除 ASCII 字元，僅保留中文用於 fuzzy 比對
_ASCII_CHARS_RE = re.compile(r"[A-Za-z0-9+\-]+")
_WHITESPACE_RE = re.compile(r"\s+")


def _normalize_cn_only(name: str) -> str:
    """僅保留中文字元（去英數、去空白），用於 fuzzy 匹配。"""
    norm = _normalize_name(name)
    norm = _ASCII_CHARS_RE.sub("", norm)
    norm = _WHITESPACE_RE.sub("", norm)
    return norm.strip()


def _parse_sku(raw: object) -> str | None:
    """將 SKU 欄位轉為純字串（去除小數點、去空白）。"""
    if raw is None:
        return None
    if isinstance(raw, float):
        raw_str = str(int(raw))
    else:
        raw_str = str(raw)
    # 去除組合包 SKU 中的括號說明（例：97890(含量杯)）
    raw_str = raw_str.split("(")[0].split("（")[0].strip()
    return raw_str if raw_str else None


def _parse_price(raw: str | None) -> tuple[int | None, str]:
    """解析價格欄，回傳 (int | None, price_display)。"""
    if not raw or not str(raw).strip():
        return None, ""
    display = str(raw).strip()
    m = _PRICE_RE.search(display)
    if m:
        try:
            price = int(m.group().replace(",", ""))
            return price, display
        except ValueError:
            pass
    logger.warning("無法解析價格：%r", display)
    return None, display


def _parse_scenarios(raw: str | None) -> tuple[str, ...]:
    """解析適用情境欄，以分隔符切割後回傳 tuple。"""
    if not raw or not str(raw).strip():
        return ()
    parts = _SCENARIO_SPLIT_RE.split(str(raw))
    return tuple(p.strip() for p in parts if p.strip())


def _parse_images_docx(
    docx_path: Path,
) -> dict[str, tuple[str, str, str]]:
    """
    解析圖片連結 docx，回傳 {原始名稱: (正規化名稱, 中文正規化, url)} dict。
    """
    doc = Document(str(docx_path))
    result: dict[str, tuple[str, str, str]] = {}
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        m = _IMAGE_RE.match(text)
        if m:
            raw_name = m.group("name").strip()
            url = m.group("url").strip().rstrip("]")
            norm = _normalize_name(raw_name)
            cn = _normalize_cn_only(raw_name)
            result[raw_name] = (norm, cn, url)
        else:
            logger.debug("無法解析圖片行：%r", text)
    return result


def _match_image(
    product_name: str,
    images: dict[str, tuple[str, str, str]],
) -> tuple[str | None, float | None]:
    """
    四層圖片匹配策略，回傳 (url | None, confidence | None)。

    1. 精確匹配（原始名稱）
    2. 正規化匹配（NFKC + 去括號）
    3. 中文子字串包含（product_cn in image_cn）
    4. fuzzy CN 匹配（threshold >= 0.85）
    """
    norm_product = _normalize_name(product_name)
    cn_product = _normalize_cn_only(product_name)

    # 第一層：精確匹配
    if product_name in images:
        return images[product_name][2], 1.0

    # 第二層：正規化匹配
    for raw_name, (norm_img, cn_img, url) in images.items():
        if norm_img == norm_product:
            return url, 0.95

    # 第三層：中文子字串包含（product_cn 至少 2 字元才觸發，防止誤匹配）
    if len(cn_product) >= 2:
        for raw_name, (norm_img, cn_img, url) in images.items():
            if cn_product in cn_img:
                ratio = len(cn_product) / max(len(cn_img), 1)
                return url, round(0.8 + ratio * 0.1, 3)

    # 第四層：fuzzy CN 匹配
    best_ratio = 0.0
    best_url: str | None = None
    for raw_name, (norm_img, cn_img, url) in images.items():
        ratio = difflib.SequenceMatcher(None, cn_product, cn_img).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_url = url

    if best_ratio >= FUZZY_THRESHOLD:
        return best_url, best_ratio

    return None, None


def _top3_candidates(
    product_name: str,
    images: dict[str, tuple[str, str, str]],
) -> tuple[str, ...]:
    """回傳 top-3 相似度的 docx 圖片名稱（以中文正規化比對）。"""
    cn_product = _normalize_cn_only(product_name)
    scored = [
        (difflib.SequenceMatcher(None, cn_product, cn_img).ratio(), raw_name)
        for raw_name, (_, cn_img, _) in images.items()
    ]
    scored.sort(reverse=True)
    return tuple(name for _, name in scored[:3])


def parse_products(catalog_path: Path, images_docx_path: Path) -> ProductsSchema:
    """
    解析產品目錄 xlsx 與圖片連結 docx，回傳標準化 ProductsSchema。

    Args:
        catalog_path: 產品目錄 xlsx 路徑
        images_docx_path: 圖片連結 docx 路徑

    Returns:
        ProductsSchema frozen dataclass 實例

    Raises:
        FileNotFoundError: 任一檔案不存在
    """
    if not catalog_path.exists():
        raise FileNotFoundError(f"產品目錄 xlsx 不存在：{catalog_path}")
    if not images_docx_path.exists():
        raise FileNotFoundError(f"圖片連結 docx 不存在：{images_docx_path}")

    sha256 = _sha256_file(catalog_path)
    generated_at = datetime.now(ZoneInfo("Asia/Taipei")).isoformat()

    meta = Meta(
        source_file="product_catalog.xlsx",
        source_sha256=sha256,
        generated_at=generated_at,
        parser_version=PARSER_VERSION,
    )

    # 解析圖片 docx
    images = _parse_images_docx(images_docx_path)

    # 解析產品目錄
    wb = openpyxl.load_workbook(catalog_path, read_only=True, data_only=True)
    ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb.active

    # sku -> Product（重複取最後一筆）
    products_dict: dict[str, Product] = {}

    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        name_raw, sku_raw, category_raw, scenario_raw, price_raw, bundle_raw, usage_raw = row

        # 跳過空列
        if not name_raw and not sku_raw:
            continue

        sku = _parse_sku(sku_raw)
        if not sku:
            logger.warning("跳過無 SKU 的產品：%r", name_raw)
            continue

        name = str(name_raw).strip() if name_raw else ""
        category = str(category_raw).strip() if category_raw else ""
        scenario = _parse_scenarios(scenario_raw)
        price, price_display = _parse_price(price_raw)
        bundle_contents = str(bundle_raw).strip() if bundle_raw and str(bundle_raw).strip() else None
        usage_note = str(usage_raw).strip() if usage_raw else ""

        # 圖片匹配
        image_url, confidence = _match_image(name, images)

        if sku in products_dict:
            logger.warning("SKU 重複 %r，取最後一筆（%r）", sku, name)

        products_dict[sku] = Product(
            sku=sku,
            name=name,
            category=category,
            scenario=scenario,
            price=price,
            price_display=price_display,
            bundle_contents=bundle_contents,
            usage_note=usage_note,
            image_url=image_url,
            image_match_confidence=confidence,
        )

    wb.close()

    # 找出未匹配的圖片
    matched_urls: set[str] = {
        p.image_url for p in products_dict.values() if p.image_url
    }
    unmatched: list[UnmatchedImage] = []
    for raw_name, (_norm, _cn, url) in images.items():
        if url not in matched_urls:
            candidates = _top3_candidates(raw_name, images)
            # 排除自身
            candidates = tuple(c for c in candidates if c != raw_name)[:3]
            unmatched.append(
                UnmatchedImage(
                    name_in_docx=raw_name,
                    url=url,
                    candidates=candidates,
                )
            )

    return ProductsSchema(
        _meta=meta,
        products=tuple(products_dict.values()),
        unmatched_images=tuple(unmatched),
    )
